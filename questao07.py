from openpyxl import load_workbook
import matplotlib.pyplot as myPlot

arquivo_excel = load_workbook(filename='chuvas_feira.xlsx')
planilha_excel = arquivo_excel['Sheet1']

categorias = []
quantidades = []
for linha in planilha_excel.iter_rows(min_row=1, max_row=12, min_col=1, max_col=2):
    mes = linha[0].value
    qnt_chuva = linha[1].value
    categorias.append(mes)
    quantidades.append(qnt_chuva)

myPlot.pie(quantidades, labels=categorias, autopct='%1.1f%%', shadow=True)

myPlot.show()