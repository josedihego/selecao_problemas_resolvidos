from openpyxl import load_workbook
import xlsxwriter

arquivo_excel = load_workbook(filename='chuvas_feira.xlsx')
planilha_excel = arquivo_excel['Sheet1']

media = 0
totalAnual = 0
for linha in planilha_excel.iter_rows(min_row=1, max_row=12, min_col=1, max_col=2):
    qnt_chuva = linha[1].value
    totalAnual = totalAnual + qnt_chuva

media = totalAnual/12

arquivo_saida_excel = xlsxwriter.Workbook('saida_chuvas.xlsx')
planilha_saida_excel  = arquivo_saida_excel.add_worksheet()
planilha_saida_excel.write('A1', 'média')
planilha_saida_excel.write('B1', media)
planilha_saida_excel.write('A2', 'Total Anual')
planilha_saida_excel.write('B2', totalAnual)
planilha_saida_excel.insert_image('C4','imagem.jpg')
arquivo_saida_excel.close()

